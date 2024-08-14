import openpyxl
import pandas as pd

def total_display(df_uploaded, excel_filename, dynamic_columns):
    # 1. OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)

    # '全体'シートを取得
    sheet_name = "全体"
    if sheet_name not in wb.sheetnames:
        print(f"'{sheet_name}'シートが存在しません。")
        return
    sheet = wb[sheet_name]

    # 2. 指定された行（ここでは2行目）にデータを挿入する準備
    insert_row_index = 2  # 転記を開始する行

    # 3. CSVデータの対象行を処理（"部門"列が空白の行のみを処理）
    for index, row in df_uploaded.iterrows():
        if pd.isna(row['部門']):
            account_name = row['Unnamed: 1']

            # 4. CSVのデータをExcelに転記
            if not pd.isna(account_name):
                new_row = [account_name]  # '勘定科目'に転記

                # 各月のデータを転記
                for col in dynamic_columns:
                    new_row.append(row[col])

                # '期間累計'を転記
                new_row.append(row['期間累計'])

                # 指定した行にデータを挿入
                for col_num, value in enumerate(new_row, start=1):
                    sheet.cell(row=insert_row_index, column=col_num, value=value)

                insert_row_index += 1

    # 変更を保存
    wb.save(excel_filename)
