import openpyxl
import pandas as pd

def process_excel_data_1(df_uploaded, excel_filename, dynamic_columns):
    
    # OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)

    # '売上高'と'売上高 計'の間にある行のインデックスを特定
    between_sales = False
    rows_to_insert_indices = []
    for index, row in df_uploaded.iterrows():
        account_name = row['Unnamed: 1']
        if account_name == '売上高':
            between_sales = True
        elif account_name == '売上高 計':
            between_sales = False
        elif between_sales and not pd.isna(account_name):
            rows_to_insert_indices.append(index)

    # アップロードされたCSVファイルの各行を処理
    for index, row in df_uploaded.iterrows():
        account_name = row['Unnamed: 1']
        if pd.isna(account_name) or index not in rows_to_insert_indices:
            continue  # 条件に当てはまらない行は無視

        # 部門名を取得し、シート名を決定
        # NaN値の場合は'全体'とする
        department = row['部門'] if not pd.isna(row['部門']) else '全体'
        sheet_name = department[:30]  # シート名の長さ制限に対応

        # 対応するシートが存在するか確認し、存在しない場合は次の行へ
        if sheet_name not in wb.sheetnames:
            continue

        # 対応するシートを取得
        sheet = wb[sheet_name]

        # 転記位置を決定（'売上高'と'売上高 計'の間のみ転記）
        insert_row = 2  # 初期値の50行の空白行の開始位置

        # 新しい行を挿入
        sheet.insert_rows(insert_row)
        new_row = [account_name, '']  # "決算書表示名"は空白とする

        # 各年月日の列の値を取得
        for col in dynamic_columns:
            new_row.append(row[col])

        # '期間累計'列の値を取得
        new_row.append(row['期間累計'])

        # 新しい行をシートに追加
        for col_idx, value in enumerate(new_row, start=1):
            sheet.cell(row=insert_row, column=col_idx, value=value)

    # 変更を保存
    wb.save(excel_filename)
