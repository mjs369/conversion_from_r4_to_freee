import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO

def app6():

    # Streamlitの設定
    st.title('試算表変換BS')
    st.markdown("""
        ### freeeからの出力手順
        1. レポートから「試算表」を選択
        2. 「エクスポート」から「CSV・PDFエクスポート」の中の「「CSV」を選択
        3. 「項目を整理する」をチェックし、「出力開始」
        4. 「エクスポート」から「出力した帳表一覧」を選択
        5. 出力したCSVのダウンロードボタンをクリック
        """)


    uploaded_file = st.file_uploader("freeeから出力した試算表(BS)をアップロードしてください", type="csv")

    if uploaded_file is not None:
        # CSVファイルを読み込む
        df = pd.read_csv(uploaded_file, encoding='cp932', header=1)

        # 数値をカンマ区切りで表示するための設定
        pd.set_option('display.float_format', '{:,.2f}'.format)

        # '資産 計'が含まれる行のインデックスを取得
        index = df[df['Unnamed: 0'] == '資産 計'].index[0]

        # '資産 計'より下の行を抽出
        new_data = df.iloc[index+1:].reset_index(drop=True)

        # 転記先の列名を変更
        new_columns = {col: f"C_{col}" for col in new_data.columns}
        new_data.rename(columns=new_columns, inplace=True)

        # '資産 計'までの行を保持
        df = df.iloc[:index+1].reset_index(drop=True)

        # 転記元の列名を変更
        df.rename(columns=lambda x: f"D_{x}", inplace=True)

        # 元のDataFrameに新しい列を追加
        result_df = pd.concat([df, new_data], axis=1)

        # カラム名を変更
        result_df.rename(columns={'D_Unnamed: 0': '借方勘定科目名'}, inplace=True)
        result_df.rename(columns={'C_Unnamed: 0': '貸方勘定科目名'}, inplace=True)

        # 一時的にExcelファイルに保存
        temp_file = BytesIO()
        result_df.to_excel(temp_file, index=False)
        temp_file.seek(0)  # ポインタを先頭に戻す

        # Excelファイルを開く
        wb = load_workbook(temp_file)
        ws = wb.active

        # 黄緑色の塗りつぶしと太文字スタイル
        highlight_fill = PatternFill(start_color="99ffcc", end_color="99ffcc", fill_type="solid")
        bold_font = Font(bold=True, name='MS Mincho')  # フォントをMS明朝に設定
        yellow_fill = PatternFill(start_color="e0ffe0", end_color="e0ffe0", fill_type="solid")
        no_fill = PatternFill(fill_type=None)  # 塗りつぶし解除のスタイル

        # 罫線のスタイルを定義
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # 列名の取得
        columns = [cell.value for cell in ws[1]]

        # 列インデックスの取得
        debit_account_col_idx = columns.index('借方勘定科目名') + 1
        credit_account_col_idx = columns.index('貸方勘定科目名') + 1

        # D_ および C_ 列のインデックスを取得
        d_columns_idx = [i + 1 for i, col in enumerate(columns) if col.startswith('D_') and not col.endswith('構成比')]
        c_columns_idx = [i + 1 for i, col in enumerate(columns) if col.startswith('C_') and not col.endswith('構成比')]
        d_ratio_columns_idx = [i + 1 for i, col in enumerate(columns) if col.startswith('D_') and col.endswith('構成比')]
        c_ratio_columns_idx = [i + 1 for i, col in enumerate(columns) if col.startswith('C_') and col.endswith('構成比')]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_index = row[0].row

            # 'D_'で始まる列のデータが全てNaNの場合
            if all(ws.cell(row=row_index, column=col).value is None for col in d_columns_idx):
                ws.cell(row=row_index, column=debit_account_col_idx).fill = highlight_fill
                ws.cell(row=row_index, column=debit_account_col_idx).font = bold_font
                for col in d_columns_idx:
                    ws.cell(row=row_index, column=col).fill = highlight_fill
                    ws.cell(row=row_index, column=col).font = bold_font

            # 'C_'で始まる列のデータが全てNaNの場合
            if all(ws.cell(row=row_index, column=col).value is None for col in c_columns_idx):
                ws.cell(row=row_index, column=credit_account_col_idx).fill = highlight_fill
                ws.cell(row=row_index, column=credit_account_col_idx).font = bold_font
                for col in c_columns_idx:
                    ws.cell(row=row_index, column=col).fill = highlight_fill
                    ws.cell(row=row_index, column=col).font = bold_font

            # '借方勘定科目名'の値が' 計'で終わっている場合
            if ws.cell(row=row_index, column=debit_account_col_idx).value and str(ws.cell(row=row_index, column=debit_account_col_idx).value).endswith(' 計'):
                ws.cell(row=row_index, column=debit_account_col_idx).fill = yellow_fill
                ws.cell(row=row_index, column=debit_account_col_idx).font = bold_font
                for col in d_columns_idx:
                    ws.cell(row=row_index, column=col).fill = yellow_fill
                    ws.cell(row=row_index, column=col).font = bold_font

            # '貸方勘定科目名'の値が' 計'で終わっている場合
            if ws.cell(row=row_index, column=credit_account_col_idx).value and str(ws.cell(row=row_index, column=credit_account_col_idx).value).endswith(' 計'):
                ws.cell(row=row_index, column=credit_account_col_idx).fill = yellow_fill
                ws.cell(row=row_index, column=credit_account_col_idx).font = bold_font
                for col in c_columns_idx:
                    ws.cell(row=row_index, column=col).fill = yellow_fill
                    ws.cell(row=row_index, column=col).font = bold_font

            # '借方勘定科目名'が空白（欠損値を含む）だった場合
            if ws.cell(row=row_index, column=debit_account_col_idx).value in [None, '']:
                ws.cell(row=row_index, column=debit_account_col_idx).fill = no_fill
                for col in d_columns_idx:
                    ws.cell(row=row_index, column=col).fill = no_fill

            # '貸方勘定科目名'が空白（欠損値を含む）だった場合
            if ws.cell(row=row_index, column=credit_account_col_idx).value in [None, '']:
                ws.cell(row=row_index, column=credit_account_col_idx).fill = no_fill
                for col in c_columns_idx:
                    ws.cell(row=row_index, column=col).fill = no_fill

        # 全てのセルに罫線を追加し、フォントをMS明朝に変更
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='MS Mincho')
                # D_ および C_ 列のセルにカンマを入れて数値を見やすくする（D_構成比とC_構成比を除く）
                if (cell.column in d_columns_idx + c_columns_idx) and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                elif (cell.column in d_ratio_columns_idx + c_ratio_columns_idx) and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

        # 列幅の自動調整（ただしデフォルトより狭くなる場合は行わない）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 列名を取得
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            current_width = ws.column_dimensions[column].width
            if adjusted_width > current_width:
                ws.column_dimensions[column].width = adjusted_width

        # 編集されたExcelファイルをバイナリストリームに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ダウンロードリンクを表示
        st.download_button(
            label="ダウンロード: 試算表BS_processed_styled.xlsx",
            data=output,
            file_name="試算表(貸借対照表).xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
