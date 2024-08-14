import openpyxl
from openpyxl.styles import Font, Border, Side

def process_excel_data_8(df_uploaded, excel_filename, dynamic_columns, uploaded_filename):
    
    # OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)

    # 列幅を調整するための関数
    def adjust_column_width(sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            current_width = sheet.column_dimensions[column_letter].width

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            # 現在の列幅より小さくならないように調整
            new_width = max(max_length + 2, current_width)
            sheet.column_dimensions[column_letter].width = new_width

    # 装飾を設定する行のリスト
    decorated_rows = ["売上高", "売上原価", "売上総損益", "販売費及び一般管理費",
                      "営業損益", "営業外利益", "営業外損失", "経常損益", "特別利益",
                      "特別損失", "税引前当期純利益"]

    # 各シートを処理
    for sheet in wb.worksheets:
        # ヘッダーの装飾を削除
        for cell in sheet[1]:
            cell.font = Font(bold=False)
            cell.border = Border()

        # 列幅を調整
        adjust_column_width(sheet)

        # 特定の行に装飾を適用
        for row in sheet.iter_rows(min_row=2):
            if row[0].value in decorated_rows:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        # 1行目にタイトルを挿入
        sheet.insert_rows(1)
        sheet['A1'] = f'{uploaded_filename} - {sheet.title}'
        sheet['A1'].font = Font(bold=True)

    # 変更を保存
    wb.save(excel_filename)

    # OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)

    # 各行の計算範囲を定義
    calc_ranges = {
        "売上高": (1, "売上高"),
        "売上原価": ("売上高", "売上原価"),
        "販売費及び一般管理費": ("売上原価", "販売費及び一般管理費"),
        "営業外利益": ("営業損益", "営業外利益"),
        "営業外損失": ("営業外利益", "営業外損失"),
        "特別利益": ("経常損益", "特別利益"),
        "特別損失": ("特別利益", "特別損失"),
    }

    # 各シートの計算を行う
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value in calc_ranges:
                start, end = calc_ranges[row[0].value]

                # 開始と終了の行を見つける
                start_row = 1 if start == 1 else None
                end_row = None
                for r in sheet.iter_rows(min_row=2, max_col=1):
                    if r[0].value == start:
                        start_row = r[0].row + 1  # 開始行は含めない
                    elif r[0].value == end:
                        end_row = r[0].row  # 終了行は含める
                        break

                if start_row is not None and end_row is not None:
                    for col_idx in range(3, sheet.max_column + 1):  # 各月および「期間累計」列
                        total = 0
                        for r in sheet.iter_rows(min_row=start_row, max_row=end_row-1, min_col=col_idx, max_col=col_idx):
                            cell_value = r[0].value
                            if cell_value is not None and isinstance(cell_value, (int, float)):
                                total += cell_value

                        # 合計を出力
                        sheet.cell(row=row[0].row, column=col_idx, value=total)

    # 変更を保存
    wb.save(excel_filename)

    # OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)
    
    def calculate_and_output(sheet, calculation_rows, output_row_name):
        row_indices = {row[0].value: row[0].row for row in sheet.iter_rows(min_row=2, max_col=1)}
        output_row_idx = row_indices.get(output_row_name)

        if output_row_idx is None:
            return

        for col_idx in range(3, sheet.max_column + 1):  # 各月および「期間累計」列
            result = 0
            for calc_row_name, sign in calculation_rows:
                row_idx = row_indices.get(calc_row_name)
                if row_idx is None:
                    continue

                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, (int, float)):
                    result += sign * cell_value

            # 結果を出力
            cell = sheet.cell(row=output_row_idx, column=col_idx, value=result)
            # 既存のフォントスタイルを保持しつつ、マイナスの場合は文字色を赤色に変更
            current_font = cell.font
            if result < 0:
                cell.font = Font(bold=current_font.bold, color="FF0000")
            else:
                cell.font = Font(bold=current_font.bold)

    # 計算と出力を行う
    for sheet in wb.worksheets:
        calculate_and_output(sheet, [("売上高", 1), ("売上原価", -1)], "売上総損益")
        calculate_and_output(sheet, [("売上総損益", 1), ("販売費及び一般管理費", -1)], "営業損益")
        calculate_and_output(sheet, [("営業損益", 1), ("営業外利益", 1), ("営業外損失", -1)], "経常損益")
        calculate_and_output(sheet, [("経常損益", 1), ("特別利益", 1), ("特別損失", -1)], "税引前当期純利益")

    # 各シートを処理
    for sheet in wb.worksheets:
        # 全てのセルを走査
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    # 数値が含まれるセルに桁区切りの数値書式を設定
                    cell.number_format = '#,##0'

    # 変更を保存
    wb.save(excel_filename)
            
    # 保存したExcelファイルのファイル名（必要に応じて変更してください）
    file_name = "月次推移_損益計算書.xlsx"