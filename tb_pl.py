import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import streamlit as st
from io import BytesIO

def app7():

    # Streamlitの設定
    st.title('試算表変換PL')
    # 説明文の追加（Markdown形式）
    st.markdown("""
        ### freeeからの出力手順
        1. レポートから「試算表」を選択
        2. 「エクスポート」から「CSV・PDFエクスポート」の中の「「CSV」を選択
        3. 「項目を整理する」をチェックし、「出力開始」
        4. 「エクスポート」から「出力した帳表一覧」を選択
        5. 出力したCSVのダウンロードボタンをクリック
        """)

    # Streamlitのファイルアップロード
    uploaded_file = st.file_uploader("freeeから出力した試算表(PL)をアップロードしてください", type="csv")

    if uploaded_file is not None:
        # データフレームの読み込み
        df = pd.read_csv(uploaded_file, encoding='cp932', header=1)

        # カラム名を変更
        df.rename(columns={'Unnamed: 0': '勘定科目名'}, inplace=True)

        # データフレームを一時的なExcelファイルに保存
        temp_excel = BytesIO()
        df.to_excel(temp_excel, index=False)
        temp_excel.seek(0)

        # Excelファイルを開く
        wb = load_workbook(temp_excel)
        ws = wb.active

        # 黄緑色の塗りつぶしと太文字スタイル
        highlight_fill = PatternFill(start_color="99ffcc", end_color="99ffcc", fill_type="solid")
        bold_font = Font(bold=True, name='MS Mincho')  # フォントをMS明朝に設定
        yellow_fill = PatternFill(start_color="e0ffe0", end_color="e0ffe0", fill_type="solid")
        no_fill = PatternFill(fill_type=None)  # 塗りつぶし解除のスタイル

        # 罫線のスタイルを定義
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # 列名の取得
        columns = [cell.value for cell in ws[1]]

        # 全ての列インデックスを取得
        all_columns_idx = [i + 1 for i in range(len(columns))]

        # ① '勘定科目名'以外の列の値が全てNaN値(または空白)の場合にhighlight_fillを摘要する
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_index = row[0].row
            account_name_value = ws.cell(row=row_index, column=columns.index('勘定科目名') + 1).value

            # '勘定科目名'以外の全ての列の値がNaN値(または空白)の場合
            if all(ws.cell(row=row_index, column=col).value in [None, ''] for col in all_columns_idx if col != columns.index('勘定科目名') + 1):
                for col in all_columns_idx:
                    ws.cell(row=row_index, column=col).fill = highlight_fill
                    ws.cell(row=row_index, column=col).font = bold_font

        # ② '勘定科目名'が特定の値だった場合にyellow_fillを摘要する
        specific_account_names = [
            '売上高 計', '商品売上原価', '製品売上原価', '売上総損益金額', '販売管理費 計',
            '営業損益金額', '営業外収益', '営業外費用', '経常損益金額', '特別利益',
            '特別損失', '税引前当期純損益金額', '当期純損益金額'
        ]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_index = row[0].row
            account_name_value = ws.cell(row=row_index, column=columns.index('勘定科目名') + 1).value

            # '勘定科目名'が特定の値の場合
            if account_name_value in specific_account_names:
                for col in all_columns_idx:
                    ws.cell(row=row_index, column=col).fill = yellow_fill
                    ws.cell(row=row_index, column=col).font = bold_font

        # 全てのセルに罫線を追加し、フォントをMS明朝に変更
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='MS Mincho')
                # 数値のセルにカンマを入れて見やすくする（最後の列を除く）
                if cell.column != ws.max_column and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        # 列幅の自動調整（ただしデフォルトより狭くなる場合は行わない）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 列名を取得
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            current_width = ws.column_dimensions[column].width
            if adjusted_width > current_width:
                ws.column_dimensions[column].width = adjusted_width

        # スタイルを適用したExcelファイルを一時的なバッファに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ダウンロードボタンを作成
        st.download_button(
            label="加工済みExcelファイルをダウンロード",
            data=output,
            file_name="試算表(損益計算書).xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
