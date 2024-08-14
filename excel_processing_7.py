import openpyxl
import pandas as pd

def process_excel_data_7(df_uploaded, excel_filename, dynamic_columns):
    
    # OpenPyXLでExcelファイルを開く
    wb = openpyxl.load_workbook(excel_filename)

    # 転記位置を探す関数
    def find_row_for_insertion(sheet, start_marker, end_marker):
        start_found = False
        for row_idx, row_data in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_data[0] == start_marker:
                start_found = True
            elif row_data[0] == end_marker and start_found:
                return row_idx
        return None

    def is_row_empty(row):
        return all(cell.value is None for cell in row)

    # '特別損失'と'税引前当期純損益金額'の間にある行のインデックスを特定
    between_inventory = False
    rows_to_insert_inventory = []
    for index, row in df_uploaded.iterrows():
        account_name = row['Unnamed: 1']
        if account_name == '特別損失':
            between_inventory = True
        elif account_name == '税引前当期純損益金額':
            between_inventory = False
        elif between_inventory and not pd.isna(account_name):
            rows_to_insert_inventory.append(index)

    # アップロードされたCSVファイルの各行を処理
    for index, row in df_uploaded.iterrows():
        account_name = row['Unnamed: 1']
        if pd.isna(account_name) or index not in rows_to_insert_inventory:
            continue  # 条件に当てはまらない行は無視

        # 部門名を取得し、シート名を決定
        department = row['部門'] if not pd.isna(row['部門']) else '全体'
        sheet_name = department[:30]  # シート名の長さ制限に対応

        # 対応するシートが存在するか確認し、存在しない場合は次の行へ
        if sheet_name not in wb.sheetnames:
            continue

        # 対応するシートを取得
        sheet = wb[sheet_name]

        # 転記位置を決定（'特別利益'と'特別損失'の間）
        insert_row = find_row_for_insertion(sheet, '特別利益', '特別損失')

        # 新しい行を挿入
        sheet.insert_rows(insert_row)
        new_row = [account_name, '']  # "決算書表示名"は空白とする

        # 各年月日の列の値を取得
        for col in dynamic_columns:
            new_row.append(row[col])

        # '期間累計'列の値を取得
        new_row.append(row['期間累計'])

        # 新しい行をシートに追加
        for col_idx, value in enumerate(new_row, start=1):
            sheet.cell(row=insert_row, column=col_idx, value=value)

    # 各シート内の空白行を削除
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows_to_delete = []

        # 空白行を検出
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            if is_row_empty(row):
                rows_to_delete.append(row_idx)

        # 空白行を削除（逆順で削除）
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx)

    # 変更を保存
    wb.save(excel_filename)