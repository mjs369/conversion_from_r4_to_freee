import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO
def app5():

    st.title("二期比較推移表作成")
    # 説明文の追加（Markdown形式）
    st.markdown("""
        ### freeeからの出力手順
        1. レポートから「月次推移」を選択
        2. 「エクスポート」から「CSV・PDFエクスポート」の中の「「CSV」を選択
        3. 「項目を整理する」をチェック「損益計算書」のみ選択し、「出力開始」
        4. 「エクスポート」から「出力した帳表一覧」を選択
        5. 出力したCSVのダウンロードボタンをクリック
        """)

    # プルダウンメニューの選択肢と対応する値
    month_options = {
        "1ヵ月分": 2,
        "2ヵ月分": 3,
        "3ヵ月分": 4,
        "4ヵ月分": 5,
        "5ヵ月分": 6,
        "6ヵ月分": 7,
        "7ヵ月分": 8,
        "8ヵ月分": 9,
        "9ヵ月分": 10,
        "10ヵ月分": 11,
        "11ヵ月分": 12,
        "12ヵ月分": 13
    }

    # プルダウンメニューの選択
    selected_option = st.selectbox("期間を選択してください", list(month_options.keys()))

    # 選択に基づいてxの値を設定
    x = month_options[selected_option]

    # ファイルアップロード
    uploaded_pre = st.file_uploader("前期の推移表CSVファイルをアップロードしてください", type="csv", key="pre")
    uploaded_now = st.file_uploader("当期の推移表CSVファイルをアップロードしてください", type="csv", key="now")

    if uploaded_pre is not None and uploaded_now is not None:
        df_pre = pd.read_csv(uploaded_pre, encoding='cp932', header=1)
        df_now = pd.read_csv(uploaded_now, encoding='cp932', header=1)

        # '期間累計'列がNaNの行を削除
        df_pre = df_pre.dropna(subset=['期間累計'])
        df_now = df_now.dropna(subset=['期間累計'])

        # Unnamed: 0を科目名に書き換える
        df_now = df_now.rename(columns={'Unnamed: 0': '科目名'})
        df_pre = df_pre.rename(columns={'Unnamed: 0': '科目名'})

        df_pre['前期累計'] = df_pre.iloc[:, 1:x].sum(axis=1)
        df_now['当期累計'] = df_now.iloc[:, 1:x].sum(axis=1)

        # 科目の位置関係を把握する
        subject_order = df_now['科目名'].tolist()

        # 挿入位置を管理するリストを作成
        new_rows = []

        # 前期の科目が今期に存在しない場合だけ処理する
        for subject in subject_order:
            if subject not in df_pre['科目名'].values:
                # 前の科目を取得
                prev_subject = subject_order[subject_order.index(subject) - 1] if subject_order.index(subject) > 0 else None

                # 次の科目を取得
                next_subject = subject_order[subject_order.index(subject) + 1] if subject_order.index(subject) < len(subject_order) - 1 else None

                # 挿入位置を決定
                inserted = False
                for i in range(len(df_pre)):
                    if prev_subject and df_pre.iloc[i]['科目名'] == prev_subject:
                        new_rows.append((i + 1, subject))
                        print(f"Inserted {subject} after {prev_subject}")
                        inserted = True
                        break
                    if next_subject and df_pre.iloc[i]['科目名'] == next_subject:
                        new_rows.append((i, subject))
                        print(f"Inserted {subject} before {next_subject}")
                        inserted = True
                        break
                if not inserted:
                    new_rows.append((df_pre.shape[0], subject))
                    print(f"Inserted {subject} at the end")

        # 挿入位置でリストをソートする
        new_rows.sort()

        # 新しい行を指定された位置に挿入
        for idx, subject in new_rows:
            new_row = pd.DataFrame({'科目名': [subject], '当期累計': [0]})
            df_pre = pd.concat([df_pre.iloc[:idx], new_row, df_pre.iloc[idx:]]).reset_index(drop=True)
            print(f"Inserted row for {subject} at index {idx}")

        # マージする前期の列を選ぶ
        df_pre_selected = df_pre[['科目名', '前期累計']]

        # 科目名をキーとしてマージ
        merged_df = df_pre_selected.merge(df_now, on='科目名', how='left')

        # NaNを処理（前期または今期に存在しない科目の処理）
        merged_df = merged_df.fillna({'前期累計': 0, '当期累計': 0})

        # 差分を計算（例として）
        merged_df['前期差額'] = merged_df['当期累計'] - merged_df['前期累計']

        merged_df.drop('期間累計', axis=1, inplace=True)

        # 現在のカラムのリストを取得
        current_columns = merged_df.columns.tolist()

        # 新しい並び順のカラムを指定
        new_columns_order = ['科目名', '前期累計', '当期累計', '前期差額'] + [col for col in current_columns if col not in ['科目名', '前期累計', '当期累計', '前期差額']]

        # 新しい並び順でデータフレームを並び替える
        merged_df = merged_df[new_columns_order]

        # 一時的にExcelファイルに保存
        excel_buffer = BytesIO()
        merged_df.to_excel(excel_buffer, index=False, engine='openpyxl')

        # Excelファイルを開く
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # 黄緑色の塗りつぶしと太文字スタイル
        highlight_fill = PatternFill(start_color="99ffcc", end_color="99ffcc", fill_type="solid")
        bold_font = Font(bold=True, name='MS Mincho')  # フォントをMS明朝に設定
        yellow_fill = PatternFill(start_color="e0ffe0", end_color="e0ffe0", fill_type="solid")
        no_fill = PatternFill(fill_type=None)  # 塗りつぶし解除のスタイル

        # 罫線のスタイルを定義
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 列名の取得
        columns = [cell.value for cell in ws[1]]

        # 全ての列インデックスを取得
        all_columns_idx = [i + 1 for i in range(len(columns))]

        # ② '勘定科目名'が特定の値だった場合にyellow_fillを摘要する
        specific_account_names = [
            '売上高 計', '商品売上原価', '製品売上原価', '売上総損益金額', '販売管理費 計',
            '営業損益金額', '営業外収益', '営業外費用', '経常損益金額', '特別利益',
            '特別損失', '税引前当期純損益金額', '当期純損益金額'
        ]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_index = row[0].row
            account_name_value = ws.cell(row=row_index, column=columns.index('科目名') + 1).value

            # '科目名'が特定の値の場合
            if account_name_value in specific_account_names:
                for col in all_columns_idx:
                    ws.cell(row=row_index, column=col).fill = yellow_fill
                    ws.cell(row=row_index, column=col).font = bold_font

        # 全てのセルに罫線を追加し、フォントをMS明朝に変更
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='MS Mincho')
                # 数値のセルにカンマを入れて見やすくする（最後の列を除く）
                if cell.column != ws.max_column and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        # 列幅の自動調整（ただしデフォルトより狭くなる場合は行わない）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 列名を取得
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            current_width = ws.column_dimensions[column].width
            if adjusted_width > current_width:
                ws.column_dimensions[column].width = adjusted_width

        # スタイルを適用したExcelファイルを保存
        styled_excel_buffer = BytesIO()
        wb.save(styled_excel_buffer)
        styled_excel_buffer.seek(0)

        # ファイルをダウンロード可能にする
        st.download_button(label="ダウンロード - スタイル適用済み推移表",
                        data=styled_excel_buffer,
                        file_name='推移表.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')